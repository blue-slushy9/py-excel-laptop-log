# We will import openpyxl as it is more suited for low-level Excel tasks,
# we also use the alias "pyx" for convenience;
import openpyxl as pyx

# Feel free to add your name if you end up contributing to this script;
print("Welcome to the Laptop Log Python script!\n")

# PLEASE RUN THIS SCRIPT WITH THE LAPTOP LOGS FOLDER AS YOUR WORKING DIRECTORY;
spreadsheet = "./Laptop Log.xlsx"

# Load the Excel file into the workbook variable, this is necessary in order
# to be able to modify the file with openpyxl;
workbook = pyx.load_workbook("./Laptop Log.xlsx")

# Define function that will delete blank rows, if needed;
def del_blank_rows(sheet):
    # We use a list to store the rows we will delete;
    rows_to_del = []
    # sheet.iter_rows returns a generator of rows, providing access to the 
    # cells in each row; our laptop information begins at row 3, so that is the min;
    # sheet.max_row simply means we iterate to the last filled-in row;
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        # Check if all cells in the row are blank;
        if all(cell.value is None for cell in row):
            # Appends the number of a blank row to our rows_to_del list;
            # row[0] accesses the first cell in current row, row[0].row gets
            # the row number of the cell, which is then added to the list;
            rows_to_del.append(row[0].row)

    # Delete the blank rows in reverse order to avoid shifting issues;
    for row_number in reversed(rows_to_del):
        # delete_rows() is a built-in openpyxl method, the final deletion step;
        sheet.delete_rows(row_number)

# Create a list of all sheets contained in this Excel file;
all_sheets = ["laptop issued", "returned", "misc."]

# del_blank_rows function will go at beginning in case user only needs to do that;
delete = input("Do you want to delete all blank rows in a sheet? [Y/n]\n")
# Use lower() to control for non-conforming user input;
delete = delete.lower()
if delete == 'y':
    which_sheet = input("On which sheet do you want to delete all blank rows (case insensitive)?\n")
    which_sheet = which_sheet.lower()
    # If the specified sheet exists...
    if which_sheet in all_sheets:
        del_blank_rows(which_sheet)
    else:
        try_again = input(f"The sheet {which_sheet} does not exist, would you like to try again? [Y/n]\n")
        try_again = sheet.lower()
        if try_again == 'y':
        



new_laptop = input("Are you adding a new laptop? [Y/n]\n")
# We use lower() to control for non-conforming user input, e.g. "y" or "N";
new_laptop = new_laptop.lower()

if new_laptop == "n":
    exist_laptop = input("Are you updating information about an existing laptop? [Y/n]\n")
    exist_laptop = exist_laptop.lower()
    if exist_laptop == "y":
        user = input("Okay, please enter the full name of the employee to whom the laptop belongs: ")

else:
    username = input("Please enter the full name of the employee to whom the laptop is being issued (case sensitive):\n")
    model = input("Please enter the laptop model, e.g. Dell E5470:\n")
    serial = input("Please enter the serial number:\n")
    laptop_name = input("Please enter the laptop name, e.g. slushy's ThinkPad:\n")
    date = input("Please enter today's date, or the date the laptop was issued; FORMAT: 23-Jan-24:\n")

    # Laptop Issued (Sheet 1) is the one we use for new laptops; 
    sheet = workbook['Laptop Issued']
    
    # Fill in the username field in Sheet 1;
    column = 'A'
    row = 

    #blank_lines = input("Do you want to delete any blank lines in the current sheet? [Y/n]\n")

    #alphabetize = input("Do you want to alphabetize the current sheet? [Y/n]\n")
